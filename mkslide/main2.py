import streamlit as st
import string
import pandas as pd
from openpyxl import load_workbook
import numpy as np # エラーチェック用に利用

def main_gui2():

 # ファイルアップロード
 uploaded_file = st.file_uploader("Excel/CSVファイルをアップロード", type=["xlsx", "xls", "csv"])
 option_examples = st.radio("フォーマット済みのCSVファイルですか？", ["Yes", "No"], index = 1, horizontal = True)

 if uploaded_file:
    # ファイルロード（data_only=Trueで計算結果の値を取得）
    try:
        wb = load_workbook(uploaded_file, data_only=True)
        ws = wb.active
    except Exception as e:
        st.error(f"Excelファイルの読み込み中にエラーが発生しました: {e}")
        st.stop()

    # --- ユーザーインターフェース ---
    st.markdown("---")  
    df_orig = pd.read_excel(uploaded_file, header=None)
    df_orig.columns = list(string.ascii_uppercase[:len(df_orig.columns)])
    df_orig.index = range(1, len(df_orig) + 1)
    st.dataframe(df_orig) 
    st.markdown("---")
    
    # ユーザーに範囲入力 (デフォルト値を設定)
    st.markdown("**各データのExcel範囲をA1:A10のように入力してください。（1列のみ対応）**")
    file_input = st.text_input("ファイル名範囲", value="A1:A10", key="k_file")
    exam_input = st.text_input("試験範囲", value="B1:B10", key="k_exam")
    face_input = st.text_input("測定面範囲", value="C1:C10", key="k_face")
    cath_input = st.text_input("正極範囲", value="D1:D10", key="k_cath")
    mesu_input = st.text_input("測定範囲", value="E1:E10", key="k_mesu")
    elec_input = st.text_input("電解液範囲", value="F1:F10", key="k_elec")
    magn_input = st.text_input("倍率範囲", value="G1:G10", key="k_magn")
    
    # 範囲をリストにまとめる
    range_inputs = {
        "ファイル名": file_input,
        "試験": exam_input,
        "測定面": face_input,
        "正極": cath_input,
        "測定": mesu_input,
        "電解液": elec_input,
        "倍率": magn_input,
    }

    # --- データの抽出と処理 ---
    
    # 範囲ごとの値を抽出する関数 (1列範囲に限定し、エラー時はNoneを返す)
    def extract_range_data(range_str):
        if not range_str:
            return None, "範囲が空です"
        try:
            # openpyxlの仕様：1列範囲でも、cは(Cell,)というタプルになる
            cells = ws[range_str]
            data_list = [c[0].value for c in cells]
            return data_list, None
        except Exception:
            # 範囲指定エラー、複数列指定、シートエラーなど
            return None, f"'{range_str}' の範囲指定が無効です。"

    if st.button("DataFrameを生成/更新"):
        extracted_data = {}
        error_messages = []

        # 1. 全範囲からデータを抽出
        for col_name, range_str in range_inputs.items():
            data_list, error = extract_range_data(range_str)
            if error:
                error_messages.append(f"列 '{col_name}': {error}")
            extracted_data[col_name] = data_list

        # エラーメッセージの表示
        if error_messages:
            st.error("以下のエラーのため、DataFrameを生成できませんでした。")
            for msg in error_messages:
                st.write(f"- {msg}")
            st.stop()
        
        # 2. リストの長さのチェック
        lengths = {name: len(lst) for name, lst in extracted_data.items() if lst is not None}
        unique_lengths = set(lengths.values())

        if len(unique_lengths) > 1:
            st.error("❌ 抽出したリストの長さが一致しません！")
            st.warning("DataFrameの列として結合するには、全てのリストが同じ長さである必要があります。")
            st.json(lengths)
            st.stop()
        
        if not unique_lengths:
            st.warning("全ての範囲が空でした。DataFrameを生成できません。")
            st.stop()

        # 3. DataFrame化して表示
        df = pd.DataFrame(extracted_data)
        st.success(f"✅ 選択範囲を結合したDataFrameを生成しました。（{list(unique_lengths)[0]}行）")
        st.dataframe(df)
